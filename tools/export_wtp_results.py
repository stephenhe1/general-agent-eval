#!/usr/bin/env python3
"""Export every WebTestPilot evaluation result into CSVs and a multi-sheet workbook.

Reads only committed run artifacts -- per-app `summary.json`, `per_bug.csv`, the SDK
transcripts (`generation/messages.jsonl`, for the authoritative `total_cost_usd`) and the
provenance audit. Nothing is hard-coded: if a run is missing, its row is simply absent.

Sheets / files produced under the output directory:

  01_run_summary   one row per (baseline, app) run, with metrics A-E and cost
  02_per_bug       one row per bug-evaluation, long format, across all runs
  03_provenance    Baseline B suite-provenance audit
  04_verdict_matrix  bug x run pivot of verdicts, for spotting per-fault agreement

Usage:
  python tools/export_wtp_results.py [--root results/wtp_headtohead] [--out <dir>]

The workbook is written only when openpyxl is importable; CSVs are always written.
"""
from __future__ import annotations

import argparse
import csv
import glob
import json
import os
import sys

BASELINES = {
    "claude-code-baseline": "A (Claude Code baseline)",
    "playwright-agents": "B (Playwright test agents)",
}

# Metrics A-E as defined for this campaign.
METRIC_COLUMNS = [
    ("total_generated_tests", "tests_generated"),
    ("clean_stable_tests", "clean_stable_tests"),
    ("clean_flaky_tests", "clean_flaky_tests"),
    ("clean_stability_rate", "A_clean_stability_rate"),
    ("num_bugs", "bugs_evaluated"),
    ("num_activated", "bugs_activated"),
    ("bug_activation_rate", "B_activation_rate"),
    ("num_caught", "bugs_caught"),
    ("conditional_fault_detection_rate", "C_conditional_detection_rate"),
    ("end_to_end_fault_detection_rate", "D_end_to_end_detection_rate"),
    ("num_incidental_failure", "incidental_failures"),
    ("num_oracle_miss", "oracle_miss"),
    ("num_not_activated", "not_activated"),
    ("num_environment_error", "environment_error"),
]


def sdk_cost(generation_dir: str) -> str:
    """Authoritative generation cost: the SDK's own total_cost_usd, last value wins.

    Deliberately not derived from token counts -- hand-rolled token math was wrong by
    2-4x earlier in this campaign.
    """
    path = os.path.join(generation_dir, "messages.jsonl")
    if not os.path.exists(path):
        return ""
    cost = ""
    for line in open(path, errors="replace"):
        if "total_cost_usd" not in line:
            continue
        try:
            rec = json.loads(line)
        except Exception:
            continue
        stack = [rec]
        while stack:
            node = stack.pop()
            if isinstance(node, dict):
                if "total_cost_usd" in node and isinstance(
                    node["total_cost_usd"], (int, float)
                ):
                    cost = f"{node['total_cost_usd']:.2f}"
                stack.extend(node.values())
            elif isinstance(node, list):
                stack.extend(node)
    return cost


def load_provenance(root: str) -> dict:
    path = os.path.join(root, "b_provenance_audit.csv")
    out = {}
    if os.path.exists(path):
        for row in csv.DictReader(open(path)):
            out[row["app"]] = row
    return out


def collect(root: str):
    runs, per_bug = [], []
    prov = load_provenance(root)

    for baseline_dir, baseline_label in BASELINES.items():
        for app_path in sorted(glob.glob(os.path.join(root, baseline_dir, "*"))):
            if not os.path.isdir(app_path):
                continue
            app = os.path.basename(app_path)
            summary_path = os.path.join(app_path, "summary.json")
            if not os.path.exists(summary_path):
                continue
            summary = json.load(open(summary_path))

            row = {"baseline": baseline_label, "app": app}
            for src, dst in METRIC_COLUMNS:
                row[dst] = summary.get(src, "")
            row["E_incidental_failure_rate"] = (
                round(summary["num_incidental_failure"] / summary["num_bugs"], 4)
                if summary.get("num_bugs")
                else ""
            )
            row["generation_cost_usd"] = sdk_cost(os.path.join(app_path, "generation"))
            if baseline_dir == "playwright-agents":
                p = prov.get(app, {})
                row["provenance_verdict"] = p.get("verdict", "")
                row["specs_stamped"] = (
                    f"{p.get('stamped','')}/{p.get('specs','')}" if p else ""
                )
            else:
                row["provenance_verdict"] = "n/a (single-agent)"
                row["specs_stamped"] = ""
            runs.append(row)

            bug_csv = os.path.join(app_path, "per_bug.csv")
            if os.path.exists(bug_csv):
                for b in csv.DictReader(open(bug_csv)):
                    per_bug.append(
                        {
                            "baseline": baseline_label,
                            "app": app,
                            "bug_name": b.get("bug_name", ""),
                            "verdict": b.get("verdict", ""),
                            "activated": b.get("activated", ""),
                            "candidate_failures": b.get("candidate_failures", ""),
                            "aligned_failures": b.get("aligned_failures", ""),
                            "detecting_test": b.get("detecting_test", ""),
                        }
                    )
    return runs, per_bug, prov


def load_validity(root: str):
    """Curated validity judgements, keyed by (baseline-letter, app).

    Unlike every other input to this export, `run_validity.csv` is hand-authored: whether a
    run is a fair measurement of its architecture is an editorial call that cannot be derived
    from the artifacts. It is kept as data rather than prose so the caveats travel with the
    numbers instead of living only in a report nobody re-reads.
    """
    path = os.path.join(root, "run_validity.csv")
    out = {}
    if os.path.exists(path):
        for row in csv.DictReader(open(path)):
            out[(row["baseline"], row["app"])] = row
    return out


def load_causes(root: str):
    """not_activated cause breakdown, keyed by (baseline-letter, app)."""
    path = os.path.join(root, "activation_causes.csv")
    rows = list(csv.DictReader(open(path))) if os.path.exists(path) else []
    by_run = {}
    for r in rows:
        if r["verdict"] != "not_activated":
            continue
        d = by_run.setdefault((r["baseline"], r["app"]), {})
        d[r["cause"]] = d.get(r["cause"], 0) + 1
    return rows, by_run


def overall_rows(runs, per_bug, prov, causes_by_run, validity_rows=()):
    """Campaign-level totals. Every figure is computed, none typed in."""

    def s(key):
        return sum(int(r[key] or 0) for r in runs)

    def block(label):
        return {"section": label, "metric": "", "value": "", "note": ""}

    def kv(section, metric, value, note=""):
        return {"section": section, "metric": metric, "value": value, "note": note}

    out = []
    a_runs = [r for r in runs if r["baseline"].startswith("A")]
    b_runs = [r for r in runs if r["baseline"].startswith("B")]

    def agg(rs, key):
        return sum(int(r[key] or 0) for r in rs)

    out.append(block("SCOPE"))
    out.append(kv("SCOPE", "runs exported", len(runs)))
    out.append(kv("SCOPE", "applications", len({r["app"] for r in runs})))
    out.append(kv("SCOPE", "bug-evaluations", len(per_bug), "one row per fault per run"))
    out.append(
        kv(
            "SCOPE",
            "distinct faults covered - A",
            agg(a_runs, "bugs_evaluated"),
            "complete benchmark" if agg(a_runs, "bugs_evaluated") == 100 else "",
        )
    )
    out.append(kv("SCOPE", "distinct faults covered - B", agg(b_runs, "bugs_evaluated")))

    for label, rs in (("BASELINE A", a_runs), ("BASELINE B", b_runs)):
        if not rs:
            continue
        ev, arm, caught = (
            agg(rs, "bugs_evaluated"),
            agg(rs, "bugs_activated"),
            agg(rs, "bugs_caught"),
        )
        out.append(block(label))
        out.append(kv(label, "tests generated", agg(rs, "tests_generated")))
        out.append(kv(label, "clean-stable tests", agg(rs, "clean_stable_tests")))
        out.append(kv(label, "flaky tests", agg(rs, "clean_flaky_tests"), "0 = fully deterministic"))
        out.append(kv(label, "environment errors", agg(rs, "environment_error")))
        out.append(kv(label, "faults evaluated", ev))
        out.append(kv(label, "faults armed", arm))
        out.append(kv(label, "B activation rate", f"{arm/ev:.1%}" if ev else "", "armed / evaluated"))
        out.append(kv(label, "faults caught", caught))
        out.append(
            kv(label, "C conditional detection", f"{caught/arm:.1%}" if arm else "", "caught / armed")
        )
        out.append(
            kv(label, "D end-to-end detection", f"{caught/ev:.1%}" if ev else "", "caught / evaluated")
        )
        out.append(kv(label, "oracle_miss", agg(rs, "oracle_miss"), "armed but undetected"))
        out.append(kv(label, "not_activated", agg(rs, "not_activated"), "never armed"))
        cost = sum(float(r["generation_cost_usd"] or 0) for r in rs)
        out.append(kv(label, "generation cost USD", f"{cost:.2f}", "SDK total_cost_usd"))

    # not_activated causes, pooled across runs.
    pooled = {}
    for d in causes_by_run.values():
        for k, v in d.items():
            pooled[k] = pooled.get(k, 0) + v
    total_na = sum(pooled.values())
    if total_na:
        out.append(block("WHY FAULTS NEVER ARMED"))
        notes = {
            "value_difference": "scenario exercised, different data used - BENCHMARK coupling",
            "coverage_gap": "no navigation call targets the required path - TOOL",
            "navigation_depth": "sessionStorage visit counter; reachable in one test - TOOL",
            "unexplained": "path navigated, no content/state requirement found - manual read",
            "indeterminate": "no condition body extractable - manual read",
        }
        for k in ("value_difference", "coverage_gap", "navigation_depth", "unexplained", "indeterminate"):
            v = pooled.get(k, 0)
            out.append(
                kv("WHY FAULTS NEVER ARMED", k, f"{v} ({v/total_na:.0%})", notes.get(k, ""))
            )
        out.append(kv("WHY FAULTS NEVER ARMED", "total not_activated", total_na))
        bench = pooled.get("value_difference", 0)
        tool = pooled.get("coverage_gap", 0) + pooled.get("navigation_depth", 0)
        unclear = pooled.get("unexplained", 0) + pooled.get("indeterminate", 0)
        out.append(
            kv("WHY FAULTS NEVER ARMED", "-> attributable to BENCHMARK",
               f"{bench} ({bench/total_na:.0%})", "literal/value coupling")
        )
        out.append(
            kv("WHY FAULTS NEVER ARMED", "-> attributable to TOOL",
               f"{tool} ({tool/total_na:.0%})", "never navigated there, or no repeat-visit flow")
        )
        out.append(
            kv("WHY FAULTS NEVER ARMED", "-> unresolved", f"{unclear} ({unclear/total_na:.0%})",
               "needs manual reading")
        )

        # Counterfactual upper bound: armed + value_difference, per baseline.
        for label, rs in (("BASELINE A", a_runs), ("BASELINE B", b_runs)):
            if not rs:
                continue
            letter = label[-1]
            vd = sum(
                d.get("value_difference", 0)
                for (bl, _), d in causes_by_run.items()
                if bl == letter
            )
            ev, arm = agg(rs, "bugs_evaluated"), agg(rs, "bugs_activated")
            out.append(
                kv(
                    "WHY FAULTS NEVER ARMED",
                    f"B' upper bound if values matched - {label}",
                    f"{(arm+vd)/ev:.1%}" if ev else "",
                    f"({arm}+{vd})/{ev} - UPPER BOUND, not an observed rate",
                )
            )

    out.append(block("SUITE PROVENANCE (B only)"))
    for app, p in sorted(prov.items()):
        out.append(
            kv(
                "SUITE PROVENANCE (B only)",
                app,
                p.get("verdict", ""),
                f"{p.get('stamped','')}/{p.get('specs','')} specs stamped, "
                f"{p.get('generator_tasks','')} generator delegations",
            )
        )

    if validity_rows:
        out.append(block("RUN VALIDITY"))
        for v in validity_rows:
            out.append(
                kv(
                    "RUN VALIDITY",
                    f"{v['baseline']} / {v['app']}",
                    v["validity"],
                    (v["cause"] + (f" [{v['cause_owner']}]" if v["cause_owner"] else ""))
                    if v["cause"]
                    else "",
                )
            )
        bad = [v for v in validity_rows if v["validity"] not in ("valid", "valid_with_note")]
        out.append(
            kv(
                "RUN VALIDITY",
                "runs needing caveat or re-run",
                f"{len(bad)} of {len(validity_rows)}",
                "; ".join(f"{v['baseline']}/{v['app']}" for v in bad),
            )
        )

    out.append(block("HEADLINE"))
    tot_arm = s("bugs_activated")
    tot_caught = s("bugs_caught")
    out.append(
        kv(
            "HEADLINE",
            "armed faults caught, all runs",
            f"{tot_caught} / {tot_arm}",
            "faults hitting the entity the test acts on are detected; "
            "faults hitting a neighbouring entity are not",
        )
    )
    return out


def counterfactual_rows(runs, causes_by_run):
    """What would happen if the suite had used the benchmark's own literals?

    Deliberately NOT a relabelling of verdicts. A `value_difference` fault never armed --
    the DOM was never mutated -- so it cannot be recorded as `caught`. Instead this projects
    the two bounds explicitly:

      OPTIMISTIC   every value-coupled fault arms AND is detected. Requires believing the
                   suite detects everything it is shown, which the observed data refutes.
      EVIDENCE-LED every value-coupled fault arms and is then detected at the conditional
                   rate actually measured on faults that did arm.

    The gap between the two columns is the point: activation and detection are independent
    problems, and fixing the benchmark's literal coupling does not fix weak assertions.
    """
    rows = []

    def agg(rs, key):
        return sum(int(r[key] or 0) for r in rs)

    observed_armed = agg(runs, "bugs_activated")
    observed_caught = agg(runs, "bugs_caught")
    observed_rate = observed_caught / observed_armed if observed_armed else 0.0

    for r in runs:
        letter = r["baseline"][0]
        vd = causes_by_run.get((letter, r["app"]), {}).get("value_difference", 0)
        ev = int(r["bugs_evaluated"] or 0)
        arm = int(r["bugs_activated"] or 0)
        caught = int(r["bugs_caught"] or 0)
        rows.append(
            {
                "baseline": r["baseline"],
                "app": r["app"],
                "faults_evaluated": ev,
                "observed_armed": arm,
                "observed_caught": caught,
                "value_difference_faults": vd,
                "cf_armed_if_values_matched": arm + vd,
                "cf_activation_rate": f"{(arm+vd)/ev:.1%}" if ev else "",
                "cf_caught_OPTIMISTIC": caught + vd,
                "cf_detection_OPTIMISTIC": f"{(caught+vd)/ev:.1%}" if ev else "",
                "cf_caught_EVIDENCE_LED": round(caught + vd * observed_rate, 1),
                "cf_detection_EVIDENCE_LED": (
                    f"{(caught + vd*observed_rate)/ev:.1%}" if ev else ""
                ),
            }
        )

    total = {
        "baseline": "TOTAL",
        "app": "",
        "faults_evaluated": agg(runs, "bugs_evaluated"),
        "observed_armed": observed_armed,
        "observed_caught": observed_caught,
        "value_difference_faults": sum(r["value_difference_faults"] for r in rows),
        "cf_armed_if_values_matched": sum(r["cf_armed_if_values_matched"] for r in rows),
        "cf_activation_rate": "",
        "cf_caught_OPTIMISTIC": sum(r["cf_caught_OPTIMISTIC"] for r in rows),
        "cf_detection_OPTIMISTIC": "",
        "cf_caught_EVIDENCE_LED": round(
            sum(r["cf_caught_EVIDENCE_LED"] for r in rows), 1
        ),
        "cf_detection_EVIDENCE_LED": "",
    }
    rows.append(total)

    note = {k: "" for k in total}
    note["baseline"] = "NOTE"
    note["app"] = (
        f"OPTIMISTIC assumes every armed fault is detected. "
        f"EVIDENCE_LED applies the measured conditional detection rate "
        f"{observed_caught}/{observed_armed} = {observed_rate:.1%}. "
        f"Neither column is an observed result; no verdict was relabelled."
    )
    rows.append(note)
    return rows


def verdict_matrix(per_bug):
    runs = sorted({(r["app"], r["baseline"]) for r in per_bug})
    cols = [f"{a} | {b.split(' ')[0]}" for a, b in runs]
    index = {}
    for r in per_bug:
        key = (r["app"], r["bug_name"])
        col = f"{r['app']} | {r['baseline'].split(' ')[0]}"
        index.setdefault(key, {})[col] = r["verdict"]
    rows = []
    for (app, bug), by_col in sorted(index.items()):
        row = {"app": app, "bug_name": bug}
        for c in cols:
            row[c] = by_col.get(c, "")
        rows.append(row)
    return rows, ["app", "bug_name"] + cols


def write_csv(path, rows, fieldnames=None):
    if not rows:
        return 0
    fieldnames = fieldnames or list(rows[0].keys())
    with open(path, "w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)
    return len(rows)


def write_workbook(path, sheets):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return False

    wb = Workbook()
    wb.remove(wb.active)
    head_fill = PatternFill("solid", fgColor="1F3864")
    head_font = Font(color="FFFFFF", bold=True)
    verdict_colors = {
        "caught": "C6EFCE",
        "oracle_miss": "FFEB9C",
        "not_activated": "F2F2F2",
        "incidental_failure": "FFC7CE",
        "environment_error": "FFC7CE",
    }

    for name, (rows, fieldnames) in sheets.items():
        ws = wb.create_sheet(name[:31])
        if not rows:
            ws["A1"] = "no data"
            continue
        fieldnames = fieldnames or list(rows[0].keys())
        ws.append(fieldnames)
        for c in ws[1]:
            c.fill, c.font = head_fill, head_font
            c.alignment = Alignment(vertical="center", wrap_text=True)
        for r in rows:
            ws.append([r.get(f, "") for f in fieldnames])

        for i, f in enumerate(fieldnames, start=1):
            longest = max([len(str(f))] + [len(str(r.get(f, ""))) for r in rows])
            ws.column_dimensions[get_column_letter(i)].width = min(max(11, longest + 2), 46)

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                colour = verdict_colors.get(str(cell.value))
                if colour:
                    cell.fill = PatternFill("solid", fgColor=colour)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb.save(path)
    return True


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--root", default="results/wtp_headtohead")
    ap.add_argument("--out", default="results/wtp_headtohead/export")
    args = ap.parse_args()

    os.makedirs(args.out, exist_ok=True)
    runs, per_bug, prov = collect(args.root)
    if not runs:
        sys.exit(f"no runs with summary.json found under {args.root}")

    matrix_rows, matrix_cols = verdict_matrix(per_bug)
    prov_rows = list(prov.values())
    cause_rows, causes_by_run = load_causes(args.root)
    validity = load_validity(args.root)

    for r in runs:
        v = validity.get((r["baseline"][0], r["app"]), {})
        r["validity"] = v.get("validity", "unreviewed")
        r["validity_cause"] = v.get("cause", "")
        r["validity_owner"] = v.get("cause_owner", "")

    # Fold the not_activated cause breakdown into each run row.
    for r in runs:
        d = causes_by_run.get((r["baseline"][0], r["app"]), {})
        for k in ("value_difference", "coverage_gap", "navigation_depth", "unexplained", "indeterminate"):
            r[f"na_{k}"] = d.get(k, 0)

    overall = overall_rows(runs, per_bug, prov, causes_by_run, list(validity.values()))
    cf_rows = counterfactual_rows(runs, causes_by_run)

    sheets = {
        "00_overall": (overall, ["section", "metric", "value", "note"]),
        "01_run_summary": (runs, list(runs[0].keys())),
        "02_per_bug": (per_bug, list(per_bug[0].keys()) if per_bug else None),
        "03_provenance": (prov_rows, list(prov_rows[0].keys()) if prov_rows else None),
        "04_verdict_matrix": (matrix_rows, matrix_cols),
        "05_activation_causes": (
            cause_rows,
            list(cause_rows[0].keys()) if cause_rows else None,
        ),
        "06_counterfactual": (cf_rows, list(cf_rows[0].keys()) if cf_rows else None),
        "07_run_validity": (
            list(validity.values()),
            list(next(iter(validity.values())).keys()) if validity else None,
        ),
    }

    for name, (rows, fields) in sheets.items():
        n = write_csv(os.path.join(args.out, f"{name}.csv"), rows, fields)
        print(f"  {name}.csv  {n} row(s)")

    xlsx = os.path.join(args.out, "webtestpilot_results.xlsx")
    if write_workbook(xlsx, sheets):
        print(f"  webtestpilot_results.xlsx  ({len(sheets)} sheets)")
    else:
        print("  (openpyxl not installed -- workbook skipped, CSVs written)")


if __name__ == "__main__":
    main()
